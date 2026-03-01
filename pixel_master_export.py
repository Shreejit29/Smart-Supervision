from openpyxl import load_workbook
from openpyxl.styles import Alignment
from io import BytesIO
import pandas as pd


def export_pixel_perfect(master_df, template_path):
    """
    master_df: Generated master supervision dataframe
    template_path: Path to original institutional format file
    """

    wb = load_workbook(template_path)
    ws = wb.active

    # ------------------------------
    # Detect Faculty Rows
    # ------------------------------

    faculty_row_map = {}

    row = 3  # Faculty starts from row 3 (based on your format)
    while ws.cell(row=row, column=2).value:

        faculty_name = ws.cell(row=row, column=2).value
        faculty_row_map[faculty_name] = row
        row += 1

    # ------------------------------
    # Detect Date Columns
    # ------------------------------

    date_col_map = {}

    col = 4  # Dates start from column 4
    while ws.cell(row=1, column=col).value:

        date_text = ws.cell(row=1, column=col).value
        date_col_map[date_text] = col
        col += 2  # Each date takes 2 columns (Session I & II)

    # ------------------------------
    # Clear Existing 1s
    # ------------------------------

    for r in faculty_row_map.values():
        for c in range(4, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None

    # ------------------------------
    # Fill Supervision Data
    # ------------------------------

    for _, row_data in master_df.iterrows():

        faculty = row_data["Name of faculty"]
        date = row_data["Date"]
        session = row_data["Session"]

        date_header = f"Date: {date}"

        if faculty not in faculty_row_map:
            continue

        if date_header not in date_col_map:
            continue

        base_col = date_col_map[date_header]

        # Session mapping
        if "I" in session:
            col_to_fill = base_col
        else:
            col_to_fill = base_col + 1

        ws.cell(
            row=faculty_row_map[faculty],
            column=col_to_fill
        ).value = 1

    # ------------------------------
    # Recalculate Totals
    # ------------------------------

    total_col = ws.max_column

    for faculty, r in faculty_row_map.items():
        count = 0
        for c in range(4, total_col):
            if ws.cell(row=r, column=c).value == 1:
                count += 1

        ws.cell(row=r, column=total_col).value = count

    # ------------------------------
    # Save to buffer
    # ------------------------------

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
